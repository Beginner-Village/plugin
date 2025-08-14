import openpyxl
import sqlite3


def iter_data(filepath: str = "doc/国内+热门海外城市列表_V2_20240312.xlsx"):
    dataframe = openpyxl.load_workbook(filepath)
    dataframe1 = dataframe.active
    if dataframe1 is None:
        raise Exception("no sheet")
    return dataframe1.iter_rows(2, dataframe1.max_row, 0, dataframe1.max_column)

def create_table(cur):
    city_table = '''
CREATE TABLE city (
    id INTEGER PRIMARY KEY,
    city_id TEXT NOT NULL,
    name TEXT NOT NULL DEFAULT '',
    f_city_name_py TEXT,
    f_city_name_pyj TEXT,
    parent_name TEXT,
    province TEXT,
    country_name TEXT NOT NULL DEFAULT '',
    city_type INTEGER NOT NULL DEFAULT 0,
    lon REAL NOT NULL DEFAULT 0,
    lat REAL NOT NULL DEFAULT 0
)'''

    index_name = 'create index idx_name on city (name, parent_name)'
    # index_f_city_name_py = 'create index idx_f_city_name_py on city (f_city_name_py)'
    # index_f_city_name_pyj = 'create index idx_f_city_name_pyj on city (f_city_name_pyj)'

    indexs = [index_name,]

    cur.execute(city_table)
    for index in indexs:
        cur.execute(index)

def save_data_paged(cur, data):
    insert_sql = "INSERT INTO city VALUES (?,?,?,?,?,?,?,?,?,?,?)"
    try:
        cur.executemany(insert_sql, data)
    except Exception as e:
        print(f"save failed: {data}")
        raise e

def save_data(cur, iter_rows):
    paged_data = []
    for idx, row in enumerate(iter_rows):
        if len(paged_data) == 1:
            save_data_paged(cur, paged_data)
            paged_data = []
        data = [idx, *[cell.value for cell in row[:10]]]
        name: str = data[2]
        parent_name: str = data[5]
        if parent_name != "" and parent_name is not None and \
            name != parent_name and name.startswith(parent_name) and \
            name != "爱丁堡市" and name != "维滕伯格" and name != "摩纳哥城":
            patched_name: str = name.replace(parent_name, "")
            patched_name = patched_name.strip()
            patched_name = patched_name.strip("（）")
            print(f"rewrite name for: {name} -> {patched_name}")
            data[2] = patched_name
        paged_data.append(data)

    if len(paged_data) > 0:
        save_data_paged(cur, paged_data)

def main():
    con = sqlite3.connect("hiagent_plugin_mojiweather/city_id.db")
    cur = con.cursor()
    create_table(cur)
    save_data(cur, iter_data())
    con.commit()
    con.close()

if __name__ == '__main__':
    main()